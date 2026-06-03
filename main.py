"""
main.py — Backend FastAPI
MZCell Rentabilidade
"""

import json
import os
import asyncio
from datetime import datetime
from typing import Optional

from fastapi import FastAPI, HTTPException, Request, UploadFile, File
import io
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

import ml_api
import parser_rentabilidade as parser
import sync_service
import auth_service
from refresh_ml_token import loop_refresh, loop_sync_diario

# ─── PATHS ───────────────────────────────────────────────────────────────────
BASE_DIR        = os.path.dirname(__file__)
CONFIG_PATH     = os.path.join(BASE_DIR, "config.json")
USUARIOS_PATH   = os.path.join(BASE_DIR, "usuarios.json")
PRODUTOS_PATH   = os.path.join(BASE_DIR, "produtos.json")
CONTAS_PATH     = os.path.join(BASE_DIR, "contas.json")
LOG_PATH        = os.path.join(BASE_DIR, "log_atividades.json")
ALERTAS_PATH    = os.path.join(BASE_DIR, "alertas.json")
DASHBOARD_PATH  = os.path.join(BASE_DIR, "dashboard.html")

# ─── APP ─────────────────────────────────────────────────────────────────────
app = FastAPI(title="MZCell Rentabilidade API", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


_syncs_em_andamento: set = set()


async def _sync_background(conta: str):
    _syncs_em_andamento.add(conta)
    try:
        await sync_service.sincronizar_conta_func(conta)
    except Exception as e:
        log_atividade(f"Erro sync background {conta}: {e}")
    finally:
        _syncs_em_andamento.discard(conta)


@app.on_event("startup")
async def startup():
    asyncio.create_task(loop_refresh())
    asyncio.create_task(loop_sync_diario())
    # Verifica e renova token imediatamente no startup
    try:
        from refresh_ml_token import verificar_e_renovar
        await verificar_e_renovar()
        print("[startup] Token verificado/renovado com sucesso")
    except Exception as e:
        print(f"[startup] Aviso token: {e}")


# ─── HELPERS ─────────────────────────────────────────────────────────────────

def load_json(path: str) -> any:
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def save_json(path: str, data: any):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def log_atividade(msg: str):
    logs = load_json(LOG_PATH) or []
    logs.append({"ts": datetime.now().isoformat(), "msg": msg})
    save_json(LOG_PATH, logs[-500:])


# ─── FRONTEND ────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def serve_dashboard():
    with open(DASHBOARD_PATH, "r", encoding="utf-8") as f:
        return f.read()


# ─── AUTH ─────────────────────────────────────────────────────────────────────

@app.post("/auth/login")
async def login(payload: dict):
    login_val = payload.get("login", "").strip().lower()
    senha     = payload.get("senha", "")
    u = auth_service.verificar_login(login_val, senha)
    if not u:
        raise HTTPException(status_code=401, detail="Login ou senha incorretos")
    log_atividade(f"Login: {u['nome']}")
    return {
        "success":         True,
        "login":           u["login"],
        "nome":            u["nome"],
        "role":            u["role"],
        "contas":          u.get("contas", []),
        "primeiro_acesso": u.get("primeiro_acesso", False),
        "email":           u.get("email", ""),
    }

@app.post("/auth/enviar-codigo")
async def enviar_codigo(payload: dict):
    email = payload.get("email", "").strip()
    login_val = payload.get("login", "").strip()
    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Email inválido")
    usuarios = auth_service.load_usuarios()
    for u in usuarios:
        if u.get("email") == email and u["login"] != login_val:
            raise HTTPException(status_code=400, detail="Este email já está em uso")
    nome = next((u["nome"] for u in usuarios if u["login"] == login_val), login_val)
    auth_service.enviar_codigo(email, nome)
    return {"success": True, "msg": f"Código enviado para {email}"}

@app.post("/auth/verificar-codigo")
async def verificar_codigo_route(payload: dict):
    email  = payload.get("email", "").strip()
    codigo = payload.get("codigo", "").strip()
    if not auth_service.verificar_codigo(email, codigo):
        raise HTTPException(status_code=400, detail="Código inválido ou expirado")
    return {"success": True}

@app.post("/auth/atualizar-senha")
async def atualizar_senha(payload: dict):
    login_val  = payload.get("login", "").strip()
    email      = payload.get("email", "").strip()
    nova_senha = payload.get("nova_senha", "")
    if len(nova_senha) < 6:
        raise HTTPException(status_code=400, detail="Senha deve ter pelo menos 6 caracteres")
    auth_service.atualizar_email_senha(login_val, email, nova_senha)
    return {"success": True, "msg": "Acesso configurado com sucesso!"}

@app.post("/auth/recuperar-senha")
async def recuperar_senha(payload: dict):
    email    = payload.get("email", "").strip()
    usuarios = auth_service.load_usuarios()
    u = next((u for u in usuarios if u.get("email") == email), None)
    if not u:
        return {"success": True, "msg": "Se o email estiver cadastrado, você receberá um código"}
    auth_service.enviar_codigo(email, u["nome"])
    return {"success": True, "msg": "Código enviado!"}

@app.post("/auth/redefinir-senha")
async def redefinir_senha(payload: dict):
    email      = payload.get("email", "").strip()
    codigo     = payload.get("codigo", "").strip()
    nova_senha = payload.get("nova_senha", "")
    if not auth_service.verificar_codigo(email, codigo):
        raise HTTPException(status_code=400, detail="Código inválido ou expirado")
    if len(nova_senha) < 6:
        raise HTTPException(status_code=400, detail="Senha deve ter pelo menos 6 caracteres")
    usuarios = auth_service.load_usuarios()
    u = next((u for u in usuarios if u.get("email") == email), None)
    if not u:
        raise HTTPException(status_code=404, detail="Email não encontrado")
    auth_service.atualizar_email_senha(u["login"], email, nova_senha)
    return {"success": True}


@app.post("/usuario/senha")
async def alterar_senha_usuario(payload: dict):
    login_val     = payload.get("login", "").strip().lower()
    senha_atual   = payload.get("senha_atual", "")
    senha_nova    = payload.get("senha_nova", "")
    senha_confirm = payload.get("senha_nova_confirmacao", "")
    if not login_val:
        raise HTTPException(status_code=400, detail="Login não informado")
    if len(senha_nova) < 6:
        raise HTTPException(status_code=400, detail="Nova senha deve ter pelo menos 6 caracteres")
    if senha_nova != senha_confirm:
        raise HTTPException(status_code=400, detail="Confirmação de senha não confere")
    u = auth_service.verificar_login(login_val, senha_atual)
    if not u:
        raise HTTPException(status_code=401, detail="Senha atual incorreta")
    auth_service.atualizar_email_senha(login_val, u.get("email", ""), senha_nova)
    log_atividade(f"Senha alterada: {login_val}")
    return {"success": True, "msg": "Senha alterada com sucesso!"}


# ─── CONFIG / CONTAS ─────────────────────────────────────────────────────────

@app.get("/config/contas")
async def get_contas():
    config = load_json(CONFIG_PATH)
    contas_info = {}
    for key, dados in config.get("contas", {}).items():
        contas_info[key] = {
            "nome": dados.get("nome"),
            "seller_id": dados.get("seller_id"),
            "ativo": dados.get("ativo", True),
            "token_ok": bool(dados.get("access_token")),
            "token_expires_at": dados.get("token_expires_at", ""),
        }
    return contas_info


class ConfigContaPayload(BaseModel):
    conta: str
    client_id: str
    client_secret: str
    redirect_uri: Optional[str] = "https://seusite.com/oauth"

@app.post("/config/conta")
async def salvar_config_conta(payload: ConfigContaPayload):
    config = load_json(CONFIG_PATH)
    if payload.conta not in config["contas"]:
        raise HTTPException(status_code=404, detail="Conta não encontrada")
    config["contas"][payload.conta]["client_id"]     = payload.client_id
    config["contas"][payload.conta]["client_secret"] = payload.client_secret
    config["contas"][payload.conta]["redirect_uri"]  = payload.redirect_uri
    save_json(CONFIG_PATH, config)
    log_atividade(f"Config salva para {payload.conta}")
    return {"success": True}


# ─── OAUTH / TOKEN ───────────────────────────────────────────────────────────

@app.get("/auth/url/{conta}")
async def gerar_url_autorizacao(conta: str):
    config = load_json(CONFIG_PATH)
    if conta not in config["contas"]:
        raise HTTPException(status_code=404, detail="Conta não encontrada")
    url = ml_api.gerar_url_autorizacao(conta)
    return {"url": url}


class AuthCodePayload(BaseModel):
    conta: str
    code: str

@app.post("/auth/autorizar")
async def autorizar(payload: AuthCodePayload):
    resultado = await ml_api.autorizar_conta(payload.conta, payload.code)
    if not resultado["success"]:
        raise HTTPException(status_code=400, detail=resultado["error"])
    log_atividade(f"Token autorizado para {payload.conta} — seller_id: {resultado.get('seller_id')}")
    return resultado


@app.post("/auth/refresh/{conta}")
async def refresh(conta: str):
    resultado = await ml_api.refresh_token(conta)
    if not resultado["success"]:
        raise HTTPException(status_code=400, detail=resultado["error"])
    return resultado


# ─── PRODUTOS ────────────────────────────────────────────────────────────────

@app.get("/produtos/{conta}")
async def listar_produtos(conta: str):
    """Lista todos os produtos ativos da conta (via API ML)."""
    resultado = await ml_api.listar_produtos(conta)
    if not resultado["success"]:
        if resultado.get("code") == 401:
            raise HTTPException(status_code=401, detail="Token expirado — faça a autorização novamente")
        raise HTTPException(status_code=400, detail=resultado["error"])
    return resultado


@app.get("/produtos/{conta}/db")
async def produtos_db(conta: str):
    """Retorna os produtos do DB local com todos os dados de rentabilidade."""
    produtos = load_json(PRODUTOS_PATH)
    conta_produtos = {k: v for k, v in produtos.items() if v.get("conta") == conta}
    return {"conta": conta, "total": len(conta_produtos), "produtos": conta_produtos}


class CustosPayload(BaseModel):
    mlb: str
    cmv_unit: float = 0.0
    frete_unit: float = 0.0
    imposto_pct: Optional[float] = None
    comissao_pct: Optional[float] = None
    roas_objetivo: float = 0.0
    rebot_unit: float = 0.0
    rebot_inicio: str = ""
    rebot_fim: str = ""
    simulador: Optional[dict] = None

@app.post("/produtos/custos")
async def salvar_custos(payload: CustosPayload):
    """Salva/atualiza custos de um produto no DB local."""
    produtos = load_json(PRODUTOS_PATH)
    if payload.mlb not in produtos:
        raise HTTPException(status_code=404, detail="Produto não encontrado no DB")
    p = produtos[payload.mlb]
    if payload.cmv_unit      != 0.0: p["cmv_unit"]      = payload.cmv_unit
    if payload.frete_unit    != 0.0: p["frete_unit"]    = payload.frete_unit
    if payload.roas_objetivo != 0.0: p["roas_objetivo"] = payload.roas_objetivo
    if payload.rebot_unit    != 0.0: p["rebot_unit"]    = payload.rebot_unit
    if payload.rebot_inicio  != "":  p["rebot_inicio"]  = payload.rebot_inicio
    if payload.rebot_fim     != "":  p["rebot_fim"]     = payload.rebot_fim
    if payload.imposto_pct   is not None: p["imposto_pct"]  = payload.imposto_pct
    if payload.comissao_pct  is not None: p["comissao_pct"] = payload.comissao_pct
    if payload.simulador     is not None: p["simulador"]    = payload.simulador
    save_json(PRODUTOS_PATH, produtos)
    log_atividade(f"Custos atualizados: {payload.mlb}")
    return {"success": True}


# ─── SYNC — Puxa dados da API ML e atualiza DB ───────────────────────────────

@app.get("/modelos/{conta}")
async def baixar_modelo_custos(conta: str):
    """Gera e retorna o modelo .xlsx de importação de custos pré-preenchido com os produtos da conta."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    cfg      = parser.load_config()
    impostos = cfg.get("impostos", {})
    aliq_com_st = impostos.get(f"{conta}_com_st", {}).get("aliquota", 0.10)
    aliq_sem_st = impostos.get(f"{conta}_sem_st", {}).get("aliquota", 0.08)

    produtos_db    = parser.load_produtos()
    produtos_conta = sorted(
        [(mlb, p) for mlb, p in produtos_db.items() if p.get("conta") == conta],
        key=lambda x: x[1].get("titulo", ""),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Importação"

    fill_amarelo = PatternFill("solid", fgColor="FFF176")
    fill_header  = PatternFill("solid", fgColor="1A1A24")
    fill_dica    = PatternFill("solid", fgColor="EEEEEE")

    COLS    = ["A","B","C","D","E","F","G","H","I","J"]
    HEADERS = ["ID do anúncio","SKU","Título","Preço de Custo",
               "Frete","Comissão","Imposto","Rebot","Rebot Início","Rebot Fim"]
    WIDTHS  = [22, 15, 46, 16, 12, 12, 12, 12, 14, 14]

    # Linha 1 — título
    ws.merge_cells("A1:J1")
    ws["A1"].value     = f"Modelo de Importação — MZ Rentabilidade · {conta}"
    ws["A1"].font      = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Linha 2 — alíquotas editáveis (B2 = COM ST, D2 = SEM ST)
    ws["A2"].value     = "Alíquota COM ST:"
    ws["A2"].font      = Font(size=10)
    ws["A2"].alignment = Alignment(vertical="center")

    ws["B2"].value         = round(aliq_com_st * 100, 4)
    ws["B2"].fill          = fill_amarelo
    ws["B2"].font          = Font(bold=True, size=10)
    ws["B2"].number_format = "0.00"
    ws["B2"].alignment     = Alignment(horizontal="center", vertical="center")

    ws["C2"].value     = "|   Alíquota SEM ST:"
    ws["C2"].font      = Font(size=10)
    ws["C2"].alignment = Alignment(vertical="center")

    ws["D2"].value         = round(aliq_sem_st * 100, 4)
    ws["D2"].fill          = fill_amarelo
    ws["D2"].font          = Font(bold=True, size=10)
    ws["D2"].number_format = "0.00"
    ws["D2"].alignment     = Alignment(horizontal="center", vertical="center")

    try:
        from openpyxl.workbook.defined_name import DefinedName
        wb.defined_names["aliquota_com_st"] = DefinedName("aliquota_com_st", attr_text="Importação!$B$2")
        wb.defined_names["aliquota_sem_st"] = DefinedName("aliquota_sem_st", attr_text="Importação!$D$2")
    except Exception:
        pass

    ws.row_dimensions[2].height = 22

    # Linha 3 — dica (antes do cabeçalho para não ser processada como dado)
    ws.merge_cells("A3:J3")
    ws["A3"].value     = "  Imposto: COM ST ou SEM ST  ·  Comissão em % (ex: 11 para 11%)  ·  Datas: DD/MM/AAAA  ·  Deixe em branco campos sem valor"
    ws["A3"].fill      = fill_dica
    ws["A3"].font      = Font(italic=True, size=9, color="9E9E9E")
    ws["A3"].alignment = Alignment(vertical="center")
    ws.row_dimensions[3].height = 16

    # Linha 4 — cabeçalhos
    for col, header in zip(COLS, HEADERS):
        c            = ws[f"{col}4"]
        c.value      = header
        c.fill       = fill_header
        c.font       = Font(bold=True, color="FFFFFF", size=10)
        c.alignment  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 22

    # Linhas 5+ — produtos (sem linha de dica entre cabeçalho e dados)
    def _num(val):
        return val if val is not None else ""

    for row_num, (mlb, p) in enumerate(produtos_conta, start=5):
        comissao_val = round((p.get("comissao_pct") or 0.11) * 100, 2)
        tem_st_val   = "COM ST" if p.get("tem_st", True) else "SEM ST"
        ws.cell(row=row_num, column=1).value  = mlb
        ws.cell(row=row_num, column=2).value  = p.get("sku", "")
        ws.cell(row=row_num, column=3).value  = p.get("titulo", "")
        ws.cell(row=row_num, column=4).value  = _num(p.get("cmv_unit"))
        ws.cell(row=row_num, column=5).value  = _num(p.get("frete_unit"))
        ws.cell(row=row_num, column=6).value  = comissao_val
        ws.cell(row=row_num, column=7).value  = tem_st_val
        ws.cell(row=row_num, column=8).value  = _num(p.get("rebot_unit"))
        ws.cell(row=row_num, column=9).value  = p.get("rebot_inicio") or ""
        ws.cell(row=row_num, column=10).value = p.get("rebot_fim") or ""

    for col_letter, width in zip(COLS, WIDTHS):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A5"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=modelo_custos_{conta}.xlsx"},
    )


@app.post("/produtos/importar-custos/{conta}")
async def importar_custos_xlsx(conta: str, file: UploadFile = File(...)):
    """
    Lê o modelo xlsx MZCell e atualiza custos no produtos.json.
    Colunas: ID do anúncio | SKU | Título | Preço de Custo | Frete | Comissão | Imposto | Rebot | Rebot Início | Rebot Fim
    B2 = alíquota COM ST (%) · D2 = alíquota SEM ST (%) — sobrepõem config.json quando preenchidos.
    """
    import openpyxl
    from datetime import datetime

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler xlsx: {e}")

    # Busca aba pelo nome
    sheet_name = None
    for name in wb.sheetnames:
        if "relat" in name.lower() or "custo" in name.lower() or "importa" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # Lê alíquotas de B2/D2 se o modelo MZCell foi usado (sobrepõem config.json)
    def _parse_aliq(val):
        try:
            v = float(val)
            return v / 100 if v > 1 else v
        except Exception:
            return None

    aliq_com_st_override = _parse_aliq(ws["B2"].value)
    aliq_sem_st_override = _parse_aliq(ws["D2"].value)

    # Detecta linha de cabeçalho procurando "ID do anúncio" ou "SKU"
    # Cada linha é varrida de forma independente; só é aceita se contiver mlb OU sku,
    # evitando que linhas de título com "custo" sejam confundidas com cabeçalho.
    header_row = None
    col_map = {}
    for row in ws.iter_rows(min_row=1, max_row=10):
        row_map = {}
        for cell in row:
            val = str(cell.value or "").lower().strip()
            if "id do" in val or ("anúncio" in val and "id" in val):
                row_map["mlb"] = cell.column
            elif val in ["frete", "frete (r$)"]:
                row_map["frete"] = cell.column
            elif "custo" in val or "cmv" in val:
                row_map["cmv"] = cell.column
            elif "sku" in val:
                row_map["sku"] = cell.column
            elif "imposto" in val:
                row_map["imposto"] = cell.column
            elif "comiss" in val:
                row_map["comissao"] = cell.column
            elif val == "rebot\n(r$/un)" or (val.startswith("rebot") and "início" not in val and "fim" not in val and "inicio" not in val):
                row_map["rebot"] = cell.column
            elif "início" in val or "inicio" in val:
                row_map["rebot_inicio"] = cell.column
            elif "fim" in val and "rebot" in row_map:
                row_map["rebot_fim"] = cell.column
        if "mlb" in row_map or "sku" in row_map:
            col_map = row_map
            header_row = cell.row
            break

    if not header_row:
        raise HTTPException(status_code=400, detail="Cabeçalho não encontrado. Verifique o formato do arquivo.")

    # Carrega produtos.json e monta índices
    produtos_db = parser.load_produtos()
    mlb_index = {mlb: mlb for mlb in produtos_db if produtos_db[mlb].get("conta") == conta}
    sku_index  = {}
    for mlb, p in produtos_db.items():
        if p.get("conta") == conta and p.get("sku"):
            sku_index[str(p["sku"]).strip().upper()] = mlb

    def get_val(row, campo):
        if campo in col_map:
            return row[col_map[campo] - 1]
        return None

    def fmt_data(d):
        if not d:
            return ""
        d = str(d).strip()
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"]:
            try:
                return datetime.strptime(d, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass
        return d

    # Alíquotas: planilha (B2/D2) tem prioridade sobre config.json
    cfg      = parser.load_config()
    impostos = cfg.get("impostos", {})
    aliq_com_st = aliq_com_st_override if aliq_com_st_override is not None \
        else impostos.get(f"{conta}_com_st", {}).get("aliquota", 0.10)
    aliq_sem_st = aliq_sem_st_override if aliq_sem_st_override is not None \
        else impostos.get(f"{conta}_sem_st", {}).get("aliquota", 0.08)

    atualizados, nao_encontrados = [], []

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(v is not None for v in row):
            continue

        mlb_val  = str(get_val(row, "mlb") or "").strip()
        sku_val  = str(get_val(row, "sku")  or "").strip().upper()
        linha_id = mlb_val or sku_val or f"linha {row_idx}"

        # Validar campos obrigatórios (apenas colunas presentes na planilha)
        campos_faltando = []
        if not (mlb_val or sku_val):
            campos_faltando.append("ID do anúncio / SKU")
        for campo, nome in [("cmv", "Preço de Custo"), ("frete", "Frete"),
                             ("comissao", "Comissão"), ("imposto", "Imposto")]:
            if campo in col_map:
                v = get_val(row, campo)
                if v is None or str(v).strip() == "":
                    campos_faltando.append(nome)
        if campos_faltando:
            nao_encontrados.append({"id": linha_id,
                                    "motivo": f"campo(s) obrigatório(s) não preenchido(s): {', '.join(campos_faltando)}"})
            continue

        mlb = mlb_index.get(mlb_val) or mlb_index.get(f"MLB{mlb_val}") or sku_index.get(sku_val)
        if not mlb:
            nao_encontrados.append({"id": mlb_val or sku_val, "motivo": "não encontrado nos produtos cadastrados"})
            continue

        frete    = float(get_val(row, "frete")    or 0)
        cmv      = float(get_val(row, "cmv")      or 0)
        comissao = float(get_val(row, "comissao")  or 0.11)
        if comissao > 1:
            comissao = comissao / 100

        imposto_str = str(get_val(row, "imposto") or "").strip().upper()
        tem_st  = imposto_str == "COM ST"
        imposto = aliq_com_st if tem_st else aliq_sem_st

        rebot     = float(get_val(row, "rebot")        or 0)
        rb_inicio = fmt_data(get_val(row, "rebot_inicio"))
        rb_fim    = fmt_data(get_val(row, "rebot_fim"))

        produtos_db[mlb].update({
            "cmv_unit":     round(cmv, 4),
            "frete_unit":   round(frete, 4),
            "comissao_pct": round(comissao, 4),
            "imposto_pct":  round(imposto, 6),
            "tem_st":       tem_st,
            "rebot_unit":   round(rebot, 4),
            "rebot_inicio": rb_inicio,
            "rebot_fim":    rb_fim,
        })
        atualizados.append(mlb_val or sku_val)

    parser.save_produtos(produtos_db)
    log_atividade(f"Importação custos {conta}: {len(atualizados)} produtos atualizados")

    return {
        "success":           True,
        "conta":             conta,
        "atualizados":       len(atualizados),
        "itens_atualizados": atualizados,
        "nao_encontrados":   nao_encontrados,
    }


@app.post("/sync/{conta}")
async def sincronizar_conta(conta: str, dias: int = 30, date_from: str = None, date_to: str = None):
    """Sincroniza todos os produtos da conta com a API do ML."""
    resultado = await sync_service.sincronizar_conta_func(
        conta, dias=dias, date_from=date_from, date_to=date_to
    )
    if not resultado["success"]:
        raise HTTPException(status_code=400, detail=resultado.get("error", "Erro no sync"))
    return resultado


@app.post("/sync/{conta}/{mlb}")
async def sincronizar_produto(conta: str, mlb: str, date_from: str = None, date_to: str = None):
    """Sincroniza um único produto da conta com a API do ML."""
    import sync_produto
    resultado = await sync_produto.sincronizar_e_retornar(conta, mlb, date_from=date_from, date_to=date_to)
    if not resultado["success"]:
        raise HTTPException(status_code=400, detail=resultado.get("error", "Erro no sync"))
    log_atividade(f"Sync produto: {mlb} ({conta}) {date_from} → {date_to}")
    return resultado


@app.post("/sync/{conta}/{mlb}/comparar")
async def sincronizar_produto_comparar(conta: str, mlb: str, date_from: str = None, date_to: str = None):
    """Sincroniza período atual e anterior em paralelo, retorna ambos para badges de variação."""
    import sync_produto
    from datetime import datetime, timedelta

    if not date_to:
        date_to = datetime.now().strftime("%Y-%m-%d")
    if not date_from:
        date_from = (datetime.now() - timedelta(days=6)).strftime("%Y-%m-%d")

    dt_from = datetime.fromisoformat(date_from)
    dt_to   = datetime.fromisoformat(date_to)
    dias    = (dt_to - dt_from).days + 1
    prev_to   = (dt_from - timedelta(days=1)).strftime("%Y-%m-%d")
    prev_from = (dt_from - timedelta(days=dias)).strftime("%Y-%m-%d")

    atual, anterior = await asyncio.gather(
        sync_produto.sincronizar_e_retornar(conta, mlb, date_from=date_from, date_to=date_to),
        sync_produto.sincronizar_e_retornar(conta, mlb, date_from=prev_from, date_to=prev_to),
    )

    if not atual["success"]:
        raise HTTPException(status_code=400, detail=atual.get("error", "Erro no sync"))

    log_atividade(f"Sync comparar: {mlb} ({conta}) {date_from}→{date_to} vs {prev_from}→{prev_to}")
    return {**atual, "anterior": anterior}


@app.get("/sync/status")
async def get_sync_status():
    """Retorna status do último sync por conta e quais estão em andamento."""
    return {
        "em_andamento": list(_syncs_em_andamento),
        "contas": sync_service.get_sync_status(),
    }


# ─── ANOTAÇÕES ────────────────────────────────────────────────────────────────

def _anotacoes_path():
    return os.path.join(BASE_DIR, "anotacoes.json")

def _load_anotacoes():
    p = _anotacoes_path()
    if not os.path.exists(p):
        return {}
    with open(p, encoding="utf-8") as f:
        return json.load(f)

def _save_anotacoes(db: dict):
    with open(_anotacoes_path(), "w", encoding="utf-8") as f:
        json.dump(db, f, indent=2, ensure_ascii=False)


@app.get("/anotacoes/{conta}/{mlb}")
async def get_anotacoes(conta: str, mlb: str):
    """Retorna TODAS as anotações do produto, ordenadas da mais recente."""
    db = _load_anotacoes()
    anotacoes = db.get(conta, {}).get(mlb, [])
    return {"success": True, "anotacoes": sorted(anotacoes, key=lambda x: x.get("id", ""), reverse=True)}


@app.post("/anotacoes/{conta}/{mlb}")
async def salvar_anotacao(conta: str, mlb: str, payload: dict):
    """Salva anotação no banco dedicado anotacoes.json."""
    from datetime import datetime as _dt
    texto = payload.get("texto", "").strip()
    if not texto:
        raise HTTPException(status_code=400, detail="Texto não pode ser vazio")

    db = _load_anotacoes()
    if conta not in db:
        db[conta] = {}
    if mlb not in db[conta]:
        db[conta][mlb] = []

    anotacao = {
        "id":        _dt.now().strftime("%Y%m%d%H%M%S%f"),
        "texto":     texto,
        "criado_em": _dt.now().strftime("%d/%m/%Y %H:%M"),
        "criado_ts": _dt.now().isoformat(),
    }
    db[conta][mlb].append(anotacao)
    _save_anotacoes(db)
    log_atividade(f"Anotação: {mlb} ({conta}) {anotacao['criado_em']}")
    return {"success": True, "anotacao": anotacao}


@app.delete("/anotacoes/{conta}/{mlb}/{anotacao_id}")
async def deletar_anotacao(conta: str, mlb: str, anotacao_id: str):
    """Remove uma anotação pelo id."""
    db = _load_anotacoes()
    lista = db.get(conta, {}).get(mlb, [])
    nova = [a for a in lista if a.get("id") != anotacao_id]
    if len(nova) == len(lista):
        raise HTTPException(status_code=404, detail="Anotação não encontrada")
    db[conta][mlb] = nova
    _save_anotacoes(db)
    return {"success": True}


# ─── RENTABILIDADE — Agrega por período ──────────────────────────────────────

@app.get("/rentabilidade/{conta}")
async def rentabilidade(conta: str, date_from: str, date_to: str):
    """
    Retorna rentabilidade agregada de todos os produtos da conta
    no período selecionado.
    """
    produtos = load_json(PRODUTOS_PATH)
    conta_produtos = {k: v for k, v in produtos.items() if v.get("conta") == conta}

    resultado = []
    for mlb, produto in conta_produtos.items():
        agregado = parser.agregar_periodo(produto, date_from, date_to)
        resultado.append({
            "mlb":       mlb,
            "titulo":    produto.get("titulo"),
            "sku":       produto.get("sku"),
            "variacao":  produto.get("variacao"),
            "custos": {
                "cmv_unit":    produto.get("cmv_unit"),
                "frete_unit":  produto.get("frete_unit"),
                "imposto_pct": produto.get("imposto_pct"),
                "comissao_pct": produto.get("comissao_pct"),
                "roas_objetivo": produto.get("roas_objetivo"),
                "rebot_unit":  produto.get("rebot_unit"),
                "simulador":   produto.get("simulador"),
            },
            "periodo": agregado,
        })

    # Ordena por receita desc
    resultado.sort(key=lambda x: x["periodo"].get("receita", 0), reverse=True)

    # KPIs globais
    kpis = {
        "receita_total":   round(sum(r["periodo"]["receita"] for r in resultado), 2),
        "margem_total":    round(sum(r["periodo"]["margem_total"] for r in resultado), 2),
        "ads_total":       round(sum(r["periodo"]["ads_total"] for r in resultado), 2),
        "unidades_total":  sum(r["periodo"]["unidades"] for r in resultado),
        "produtos_ativos": len(resultado),
    }
    kpis["margem_pct_global"] = round(kpis["margem_total"] / kpis["receita_total"], 6) if kpis["receita_total"] > 0 else 0
    kpis["tacos_global"]      = round(kpis["ads_total"] / kpis["receita_total"], 6) if kpis["receita_total"] > 0 else 0

    return {
        "conta": conta,
        "periodo": {"de": date_from, "ate": date_to},
        "kpis": kpis,
        "produtos": resultado,
    }


# ─── LOG / ALERTAS ───────────────────────────────────────────────────────────

@app.get("/log")
async def get_log(limit: int = 100):
    logs = load_json(LOG_PATH) or []
    return {"logs": logs[-limit:]}

@app.get("/alertas")
async def get_alertas():
    return {"alertas": load_json(ALERTAS_PATH) or []}


@app.get("/sistema/versao")
async def get_versao():
    try:
        v = open(os.path.join(BASE_DIR, "version.txt"), encoding="utf-8").read().strip()
    except FileNotFoundError:
        v = "?"
    return {"versao": v}


@app.get("/sistema/houve-atualizacao")
async def houve_atualizacao_endpoint():
    version_path = os.path.join(BASE_DIR, "version.txt")
    anterior_path = os.path.join(BASE_DIR, "version_anterior.txt")
    try:
        versao_atual = open(version_path, encoding="utf-8").read().strip()
    except FileNotFoundError:
        return {"houve": False}
    versao_anterior = None
    if os.path.exists(anterior_path):
        versao_anterior = open(anterior_path, encoding="utf-8").read().strip()
    if versao_anterior != versao_atual:
        open(anterior_path, "w", encoding="utf-8").write(versao_atual)
        return {"houve": True, "versao": versao_atual}
    return {"houve": False}


if __name__ == "__main__":
    import subprocess, time, uvicorn
    try:
        saida = subprocess.check_output("netstat -ano | findstr :8001", shell=True).decode()
        for linha in saida.strip().split('\n'):
            col = linha.strip().split()
            pid = col[-1] if col else ""
            if pid.isdigit() and pid != '0':
                subprocess.run(f"taskkill /PID {pid} /F", shell=True, capture_output=True)
        time.sleep(0.5)
    except Exception:
        pass
    config = load_json(CONFIG_PATH)
    host = config.get("servidor", {}).get("host", "127.0.0.1")
    port = config.get("servidor", {}).get("port", 8001)
    uvicorn.run("main:app", host=host, port=port, reload=False)
